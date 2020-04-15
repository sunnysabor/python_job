from openpyxl import Workbook
from openpyxl import load_workbook

# 读取文件
# file = '/Users/jerry/Downloads/佛山.xlsx'
# def getDataOnly(file):
#     wb = load_workbook(file)
#     ws = wb.worksheets[0]
#     for row in ws.values:
#         for value in row:
#             print(value)
#
# getDataOnly(file)


wb = Workbook()
ws = wb.active()

for row in ws.iter_rows(min_row=1, max_row=10, max_col=10):
    for cell in row:
        cell.value
wb.save("tmp.xlsx")
